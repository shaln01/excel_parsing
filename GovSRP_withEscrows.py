# import pandas as  pd
# import numpy as np
import copy
import sys
import json
sheet_name="Govt SRP"
# df = pd.read_excel("C:\\Users\\ShailendraSingh\\Desktop\\ameri.xlsx",sheet_name=sheet_name)

# import module
import openpyxl

n = 13
# sh = wrkbk.active
def govSrp(wrkbk,version):
    sheets = wrkbk.sheetnames
    sheet_name = 'Govt SRP'
    # sh = wrkbk[sheets[n]]
    sh = wrkbk[sheet_name]
    version = version
    withEscrow =True
    provider =sheet_name

    bold_flag = False
    bold_indx = -1
    starting_row_indx =-1
    col_list =[]
    row_indxs =[]
    ending_col_indx =-1
    ending_row_index =-1,
    row_dict={
            "start":'',
            "end":""
        }
    table_headings=[]
    col_indxs = []

    def find_min_max(value):
        val = value.replace(',','')
        loan_dict ={
            "min":0,
            "max":0
        }
        if '<=' in val:
            data_min, data_max = val.split('<=')

            loan_dict["min"] = 0
            loan_dict["max"] = int(data_max)
        if '-' in val:
            data_min, data_max = val.split('-')

            loan_dict["min"] = int(data_min)
            loan_dict["max"] = int(data_max)
        if '>' in val:
            data,data_min = val.split('>')

            loan_dict["min"] = int(data_min)
            loan_dict["max"] = sys.maxsize
        return loan_dict

    def cal_term(value):
        product_info_dict={
            "productType":'',
            "term":[]
        }
        prod_packing = value.split(" ")
        terms=prod_packing[1]
        term_val = terms.split("/")
        # print('here is what you looking for',term_val)
        product_info_dict["term"]= term_val
        product_info_dict["term"]  = list(map(int, product_info_dict["term"]))
        product_info_dict["productType"] = prod_packing[-1]
        return product_info_dict
    for i in range(1, sh.max_row + 1):
        col_dict = {
            "start": '',
            "end": ""
        }
        for j in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=i, column=j)
            # print(cell_obj.value, end=" ")
            # print("new val",((sh.cell(row =bold_indx+ 2, column = j)).font.b))
            if cell_obj.font.b and cell_obj.value!= None:
                bold_indx = i
                bold_flag =True
                # print("Bold===>",cell_obj.value, bold_indx)
                table_headings.append(copy.deepcopy(cell_obj.value))
            if i == bold_indx + 2 and bold_flag == True and ((sh.cell(row =bold_indx+ 2, column = j)).value)!=None and  ((sh.cell(row =bold_indx+ 2, column = j)).font.b) == False:
                # print(sh.cell(row=i, column=j).value,'real values')
                # print(i)
                row_dict["start"] = copy.deepcopy(i)
                col_list.append(copy.deepcopy(j))
                # print(col_list)
        if i>=bold_indx +2 and bold_flag == True and sh.cell(row=i, column=col_list[0]).value ==None:
            # print(sh.cell(row=i, column=col_list[0]) )
            row_dict["end"] = (i-1)
            col_dict["start"] = col_list[0]
            col_dict["end"] = col_list[-1]
            col_indxs.append(copy.deepcopy(col_dict))
            col_list = []
            row_indxs.append(copy.deepcopy(row_dict))
            row_dict={

            }
            col_dict= {

            }
            bold_flag =False
    table_headings.pop(0)
    # print(row_indxs)
    # print(col_indxs)
    # print(table_headings)
    json_list = []
    main_json_list = []
    json_list_obj ={
        "version": version,
        "withEscrow":withEscrow,
        "provider":provider,
        "product":"",
        "state":"",
        "term":[],
        "productType":"",
        "values":[
        ]
    }

    data_value={
                "minLoan":0,
                "maxLoan":9999999999999,
                "srp": 1
            }
    for index, value in enumerate(table_headings) :
        row_val = row_indxs[index]
        col_val = col_indxs[index]
        table_heading = table_headings[index]

        for row_range  in range(row_val['start'],row_val['end']+1):
            # print(table_heading)
            for col_range in range(col_val['start'],col_val['end']+1):
                min_amt=''
                if row_range > row_val['start']:
                    state_val = (sh.cell(row=row_range, column=col_val['start'])).value
                    json_list_obj["state"] =state_val
                    loan_amt=(sh.cell(row=row_val["start"], column=col_range)).value
                    loan_amount= find_min_max((sh.cell(row=row_val["start"], column=col_range)).value)
                    product_info = cal_term(table_heading)
                    json_list_obj["product"] = table_heading
                    json_list_obj["term"]= product_info["term"]
                    json_list_obj["productType"] = product_info["productType"]
                    data_value["minLoan"]=copy.deepcopy(loan_amount["min"])
                    data_value["maxLoan"]=copy.deepcopy(loan_amount["max"])
                if row_range > row_val['start'] and col_range > col_val['start']:
                    cell_obj = sh.cell(row=row_range, column=col_range)
                    data_value["srp"] =cell_obj.value
                    json_list_obj['values'].append(copy.deepcopy(data_value))
            if row_range > row_val['start']:
                main_json_list.append(copy.deepcopy(json_list_obj))
                json_list_obj = {
                                "version": version,
                                "withEscrow": withEscrow,
                                 "provider": provider,
                                "product":"",
                                "term": [],
                                "productType": "",
                                "state":"",
                                "values":[
                                ]
                        }
    print("here is what u looking for length of the array isn't it...!" ,len(main_json_list))
    json_object = json.dumps(main_json_list, indent=4)
    with open("GovtSRP_withEscrows.json","w") as j_file:
        print("JSON file is generated Happppilyy...!")
        j_file.write(json_object)
    return (main_json_list)
    #
    #
    #
