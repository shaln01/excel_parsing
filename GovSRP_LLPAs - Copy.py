# import pandas as  pd
# import numpy as np
import copy
import sys
import json

# df = pd.read_excel("C:\\Users\\ShailendraSingh\\Desktop\\ameri.xlsx",sheet_name=sheet_name)

# import module
import openpyxl

wrkbk = openpyxl.load_workbook("C:\\Users\\ShailendraSingh\\Desktop\\ameri.xlsx")
n = 13
# sh = wrkbk.active
sheets = wrkbk.sheetnames
sheet_name = 'Government LLPAs'
# sh = wrkbk[sheets[n]]
sh = wrkbk[sheet_name]
version = 1
withEscrow =True
provider =sheet_name

bold_flag = False
bold_indx = -1
starting_row_indx =-1
col_list =[2,13]
main_col =2
main_heading_found =False
row_arr=[]
# row_indxs =[{
#         "start":14,
#         "end":23
#     },
#     {
#         "start":35,
#         "end":65
#     }]
row_indxs =[ ]
ending_col_indx =-1
ending_row_index =-1,
row_dict={
        "start":'',
        "end":""
    }
state_row_dict={
        "start":'',
        "end":""
    }
state_tire_heading_found =False
state_row_arr=[]

state_tire_col =1
product_indx_dict ={
    "name" :"",
    "start": '',
    "end": ""
}
state_row_indxs = []
stateTireGroup=[

]
product_start_indx=[]
product_list =[]
product_end_indx =[]
product_va_fha_usda_indx = -1

product_value = []
product_indx =[]
table_headings=[]
# col_indxs =[{'start': 2, 'end': 8},{'start': 2, 'end': 8}]
col_indxs =[]
def find_min_max(value):
    print("FICO VALLLLL ",value)
    val = value.replace(' ','')
    val = value.replace(',','')
    val = val.replace('%','')
    val = val.replace('$','')
    val = val.replace('K','000')
    loan_dict ={
        "min":0,
        "max":0
    }
    if '≥' in val:
        data, data_min = val.split('≥')
        loan_dict["min"] = float(data_min)
        loan_dict["max"] = sys.maxsize
    if '<=' in val :
        data_min, data_max = val.split('<=')

        loan_dict["min"] = 0
        loan_dict["max"] = float(data_max)
    if '=<' in val:
        data_min, data_max = val.split('=<')

        loan_dict["min"] = 0
        loan_dict["max"] = float(data_max)


    if '-' in val:
        data_min, data_max = val.split('-')
        data_min = data_min.replace('>','')
        loan_dict["min"] = float(data_min)
        loan_dict["max"] = float(data_max)
    if '>' in val:
        data,data_min = val.split('>')
        if '-' in data_min:
            min_data, max_data = val.split('-')
            min_data = min_data.replace('>','')
            loan_dict["min"] = float(min_data)
            loan_dict["max"] = float(max_data)
        elif "=" in data_min:
            data ,min_data = val.split('=')
            loan_dict["min"]= float(min_data)
            loan_dict["max"] = sys.maxsize
        elif '-' not in data_min:
            loan_dict["min"] = float(data_min)
            loan_dict["max"] = sys.maxsize


    return loan_dict

def cal_term(value):
    product_info_dict={
        "productType":'',
        "term":[]
    }
    prod_packing = value.split(" ")
    terms=prod_packing[1]
    term_val = terms.split("/")
    # print('here is what you looking for',term_val)
    product_info_dict["term"]= term_val
    # product_info_dict["term"]  = list(map(int, product_info_dict["term"]))
    product_info_dict["productType"] = prod_packing[-1]
    return product_info_dict
for i in range(1, sh.max_row + 1):
    col_dict = {
        "start": '',
        "end": ""
    }
    for j in range(1, sh.max_column + 1):
        cell_obj = sh.cell(row=i, column=j)
        # print(cell_obj.value, end=" ")
        # print(sh.cell(row=i, column=col_list[0]).value)
        # print("new val",((sh.cell(row =bold_indx+ 2, column = j)).font.b))
        if cell_obj.font.b and cell_obj.value!= None:
            bold_indx = i
            bold_flag =True
            # print("Bold===>",cell_obj.value, bold_indx)
            table_headings.append(copy.deepcopy(cell_obj.value))
        if i== bold_indx +1 and bold_flag ==True:
            val_found =False
            tire_found = False
            if ((sh.cell(row =bold_indx+ 1, column = j)).value) in ['FHA','VA','USDA']:
                val_found =True
                product_list.append(copy.deepcopy((sh.cell(row =bold_indx+ 1, column = j)).value))
                product_start_indx.append(j)

                product_value.append(((sh.cell(row =bold_indx+ 1, column = j)).value))
            if ((sh.cell(row=bold_indx + 1, column=j)).value) in ['FHA', 'VA', 'USDA'] and val_found == True:
                product_end_indx.append(j-1)
                print("yaay")
        if i == bold_indx + 2 and ((sh.cell(row =bold_indx+ 2, column = j)).value)=='FICO':
            main_col =j
            main_heading_found =True

        if main_heading_found ==True and ((sh.cell(row =i, column = main_col)).value)== None:
            main_heading_found =False
            row_dict["start"] = copy.deepcopy(row_arr[0])
            row_dict["end"] = copy.deepcopy(row_arr[-1])
            row_arr=[]
            row_indxs.append(copy.deepcopy(row_dict))
            row_dict = {

            }
        if i == bold_indx+1 and (sh.cell(row =bold_indx+1, column = j)).value=='Tier':
            state_tire_col =j
            state_tire_heading_found =True
            print( "array of state")


        if state_tire_heading_found ==True and ((sh.cell(row =i, column = state_tire_col)).value)== None:
            state_tire_heading_found =False
            state_row_dict["start"] = copy.deepcopy(state_row_arr[0])
            state_row_dict["end"] = copy.deepcopy(state_row_arr[-1])
            state_row_arr=[]
            state_row_indxs.append(copy.deepcopy(state_row_dict))
            row_dict = {

            }
        if i>=bold_indx +1 and state_tire_heading_found== True and ((sh.cell(row =i, column = state_tire_col)).value)!=None and  ((sh.cell(row =bold_indx+ 1, column = 2)).font.b) == False and state_tire_heading_found ==True:
            print('FICO',i)
            state_row_arr.append(copy.deepcopy(i))
        if i>=bold_indx +2 and main_heading_found== True and ((sh.cell(row =i, column = main_col)).value)!=None and  ((sh.cell(row =bold_indx+ 2, column = 2)).font.b) == False and main_heading_found ==True:
            # print('FICO',i)
            row_arr.append(copy.deepcopy(i))
        if i == bold_indx + 2 and bold_flag == True and ((sh.cell(row =bold_indx+ 2, column = j)).value)!=None and  ((sh.cell(row =bold_indx+ 2, column = j)).font.b) == False:
            # print(sh.cell(row=i, column=j).value,'real values')
            # print(i)
            # row_dict["start"] = copy.deepcopy(i)
            col_list.append(copy.deepcopy(j))
            # print(col_list)
        # if ((sh.cell(row=bold_indx + 1, column=j)).value) in ['Tier']:
        #     tire_found = True
        #     if tire_found and i == bold_indx + 2:
        #         pass
    if i>=bold_indx +2 and bold_flag == True and sh.cell(row=i, column=col_list[0]).value ==None:
        # print(sh.cell(row=i, column=col_list[0]) )
        # row_dict["end"] = (i-1)
        col_dict["start"] = col_list[0]
        col_dict["end"] = col_list[-1]
        if sh.cell(row=bold_indx +2, column=col_list[0]).value =="FICO":
            product_end_indx.pop(0)
            product_end_indx.append(copy.deepcopy(col_list[-1]))
        col_indxs.append(copy.deepcopy(col_dict))
        col_list = []

        col_dict= {

        }
        bold_flag =False
table_headings.pop(0)
state_table=table_headings.pop(1)
table_headings.pop()
table_headings.pop()
# row_indxs.pop(1)
state_rows=[]
state_tire_col_indxs=col_indxs.pop(1)

#for rows
state_tier_list=[]
state_col = state_tire_col_indxs
row_col = state_row_indxs[0]
state_obj_dict={
    "tier":'',

}
for row_range in range(row_col['start']+1, row_col['end'] + 1):
        state_obj_dict["tier"] = (sh.cell(row=row_range, column=state_col['start'],)).value
        state_obj_dict["stateGroup"]= (((sh.cell(row=row_range, column=state_col['end'],)).value)).replace(' ','').split(",")
        state_tier_list.append(copy.deepcopy(state_obj_dict))
        state_obj_dict={

        }

print(product_end_indx)
print("state cols",state_tier_list,state_tire_heading_found)
# print(row_indxs)
# print(col_indxs)
print(table_headings)
print(row_indxs)
print(col_indxs)
json_list = []
main_json_list = []
json_list_obj ={
    "version": version,
    "withEscrow":withEscrow,
    "provider":provider,
    "product":"",
    #
    # "term":[],
    # "productType":"",

    "values":[
    ]
}

data_value={
            "minLoan":0,
            "maxLoan":9999999999999,
        }
tire_val =100
for index, value in enumerate(table_headings) :
    row_val = row_indxs[index]
    col_val = col_indxs[index]
    table_heading = table_headings[index]
    fico_col =1

    for row_range  in range(row_val['start'],row_val['end']+1):
        # print(table_heading)

        for col_range in range(col_val['start'],col_val['end']+1):
            min_amt=''
            if (sh.cell(row=row_range, column=col_range)).value == 'FICO':
                fico_found = True
                fico_col = col_range
                print("FICO", col_range)
            if row_range > row_val['start']:

                if col_range == fico_col:
                    pass
                    # print("YESS ", fico_col)
                state_val = (sh.cell(row=row_range, column=col_val['start'])).value
                check1 ="fico"
                check2 ="tires"
                if  check1.casefold() not in table_heading.casefold() and check2.casefold() not in table_heading.casefold():
                    state_val = (sh.cell(row=row_range, column=col_val['start']+1)).value
                    tier_loc_val = (sh.cell(row=row_range, column=col_val['start'])).value
                    if (sh.cell(row=row_range, column=col_range)).value == 'FICO':
                        fico_found = True
                        fico_col = col_range
                        print("FICO", col_range)

                    if tier_loc_val != None :
                        tier_val = tier_loc_val
                        json_list_obj["tier"] = tier_val
                    if tier_loc_val ==None:
                        json_list_obj["tier"] = tier_val
                    for tire_indx ,val_tier in enumerate(state_tier_list):
                        if json_list_obj['tier'] == val_tier['tier'] :
                            json_list_obj["stateGroup"] = val_tier['stateGroup']

                    ltv = find_min_max(str(sh.cell(row=row_val["start"], column=col_range).value))
                    # product_info = cal_term(table_heading)
                    json_list_obj["product"] = table_heading
                    if state_val != None :
                        fico_val = state_val
                        fico = find_min_max(fico_val)
                        json_list_obj["minFico"] = fico['min']
                        if fico['max']==sys.maxsize:
                            json_list_obj["maxFico"] = 1000
                        elif fico['max']!=sys.maxsize:
                            json_list_obj["maxFico"] = fico['max']
                    if state_val ==None:
                        fico = find_min_max(fico_val)
                        json_list_obj["minFico"] = fico['min']
                        if fico['max']==sys.maxsize:
                            json_list_obj["maxFico"] = 1000
                        elif fico['max']!=sys.maxsize:
                            json_list_obj["maxFico"] = fico['max']

                    # json_list_obj["term"]= product_info["term"]
                    json_list_obj["productType"] =table_heading
                    data_value["minLtv"] = copy.deepcopy(ltv["min"])
                    data_value["maxLTV"] = copy.deepcopy(ltv["max"])
                    data_value["minCltv"] = copy.deepcopy(ltv["min"])
                    data_value["maxClTV"] = copy.deepcopy(ltv["max"])
                elif check1.casefold() in table_heading.casefold() and check2.casefold() not in table_heading.casefold():

                    loan_amount = find_min_max(str(sh.cell(row=row_val["start"], column=col_range).value))
                    # product_info = cal_term(table_heading)
                    json_list_obj["product"] = table_heading
                    # json_list_obj["term"]= product_info["term"]
                    # json_list_obj["productType"] = product_info["productType"]
                    data_value["minLoan"] = copy.deepcopy(loan_amount["min"])
                    data_value["maxLoan"] = copy.deepcopy(loan_amount["max"])



                    if state_val != None :



                        fico_val = state_val
                        fico = find_min_max(fico_val)
                        json_list_obj["minFico"] = fico['min']
                        if fico['max']==sys.maxsize:
                            json_list_obj["maxFico"] = 1000
                        elif fico['max']!=sys.maxsize:
                            json_list_obj["maxFico"] = fico['max']
                    if state_val ==None:
                        fico = find_min_max(fico_val)
                        json_list_obj["minFico"] = fico['min']
                        if fico['max']==sys.maxsize:
                            json_list_obj["maxFico"] = 1000
                        elif fico['max']!=sys.maxsize:
                            json_list_obj["maxFico"] = fico['max']


                    # if col_range in product_start_indx:
                    #     if product_va_fha_usda_indx < len(product_list) -1:
                    #         product_va_fha_usda_indx = product_va_fha_usda_indx +1
                    #     print(product_va_fha_usda_indx,"VA")
                    #     for end_ind in product_end_indx:
                    #         if col_range <= end_ind:
                    #             # if product_list[product_va_fha_usda_indx] in product_list:
                    #
                    #             print(product_list[product_va_fha_usda_indx],'isVA')
                    #             json_list_obj["productTyp"].append(copy.deepcopy(product_list[product_va_fha_usda_indx]))

                # print("Tire",(sh.cell(row=40, column=col_val['start'])).value)
                loan_amt=(sh.cell(row=row_val["start"], column=col_range)).value

            if row_range > row_val['start'] and col_range > fico_col:

                cell_obj = sh.cell(row=row_range, column=col_range)
                if cell_obj.value != None:
                    if check1.casefold() not in table_heading.casefold() and check2.casefold() not in table_heading.casefold():
                        data_value["llpaValue"] =cell_obj.value
                    elif  check1.casefold()  in table_heading.casefold() and check2.casefold() not in table_heading.casefold():
                        if col_range >= product_start_indx[0] and col_range<=product_end_indx[0]:
                            data_value["productType"]=product_list[0]
                        elif col_range >= product_start_indx[1] and col_range<=product_end_indx[1]:
                            data_value["productType"]=product_list[1]
                        elif col_range >= product_start_indx[2] and col_range<=product_end_indx[2]:
                            data_value["productType"]=product_list[2]
                        data_value["llpaValue"] =cell_obj.value
                else:
                    pass
                json_list_obj['values'].append(copy.deepcopy(data_value))
                data_value={

                }
        if row_range > row_val['start']:

            main_json_list.append(copy.deepcopy(json_list_obj))
            json_list_obj = {
                            "version": version,
                            "withEscrow": withEscrow,
                             "provider": provider,
                            "product":"",
                            # "term": [],
                            # "productType": "",
                            # "state":"",

                            "values":[
                            ]
                    }
print("here is what u looking for length of the array isn't it...!" ,len(main_json_list))
print(product_list)
print(product_start_indx)
print(product_end_indx)
json_object = json.dumps(main_json_list, indent=4)
with open("GovtLLPAs.json","w") as j_file:
    print("JSON file is generated Happppilyy...!")
    j_file.write(json_object)
#
#
#
