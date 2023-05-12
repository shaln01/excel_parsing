# import pandas as  pd
# import numpy as np
import copy
import sys
import json
sheet_name="Govt SRP"
# df = pd.read_excel("C:\\Users\\ShailendraSingh\\Desktop\\ameri.xlsx",sheet_name=sheet_name)

# import module
import openpyxl

# wrkbk = openpyxl.load_workbook("C:\\Users\\ShailendraSingh\\Desktop\\ameri.xlsx")
n = 13
# sh = wrkbk.active
def portfolioExpressLlpa(wrkbk,version):
    sheets = wrkbk.sheetnames
    sheet_name = 'Portfolio Express LLPAs'
    # sh = wrkbk[sheets[n]]
    sh = wrkbk[sheet_name]
    version = version
    withEscrow =True
    provider =sheet_name

    bold_flag = False
    bold_indx = -1
    starting_row_indx =-1
    col_list =[]
    row_indxs =[]
    ending_col_indx =-1
    ending_row_index =-1,
    row_dict={
            "start":'',
            "end":""
        }
    table_headings=[]
    col_indxs = []

    def find_min_max(value):
        val = value.replace(',', '')

        loan_dict ={
            "min":0,
            "max":0
        }
        if '≤' in val :
            data_min, data_max = val.split('≤')

            loan_dict["min"] = float(0)
            loan_dict["max"] = float(data_max)

        if '≥' in val:
            data, data_min = val.split('≥')
            loan_dict["min"] =float(data_min)
            loan_dict["max"] =sys.maxsize
        if '<=' in val:
            data_min, data_max = val.split('<=')

            loan_dict["min"] = 0
            loan_dict["max"] = float(data_max)
        if '-' in val:
            data_min, data_max = val.split('-')
            loan_dict["min"] = float(data_min)
            loan_dict["max"] = float(data_max)
        # if '>' in val:
        #     data,data_min = val.split('>')
        #
        #     loan_dict["min"] = float(data_min)
        #     loan_dict["max"] = sys.maxsize
        return loan_dict

    def find_min_max2(value):
        val1 = value.replace('mm', '')
        val = val1.replace('>','')
        print(val)


        loan_dict ={
            "min":0,
            "max":0
        }
        if '≥' in val:
            data, data_min = val.split('≥')
            loan_dict["min"] =float(data_min)
            loan_dict["max"] =sys.maxsize
        if '<=' in val:
            data_min, data_max = val.split('<=')

            loan_dict["min"] = 0
            loan_dict["max"] = float(data_max)
        if '-' in val:
            data_min, data_max = val.split('-')
            if '.' in data_min or '.' in data_max:
                loan_dict["min"] = float(data_min) * 1000000
                loan_dict["max"] = float(data_max) * 1000000
            else:
                loan_dict["min"] = float(data_min)
                loan_dict["max"] = float(data_max)
        # if '>' in val:
        #     data,data_min = val.split('>')
        #
        #     loan_dict["min"] = float(data_min)
        #     loan_dict["max"] = sys.maxsize
        return loan_dict

    def cal_term(value):
        product_info_dict={
            "productType":'',
            "term":[]
        }
        prod_packing = value.split(" ")
        terms=prod_packing[1]
        term_val = terms.split("/")
        # print('here is what you looking for',term_val)
        product_info_dict["term"]= term_val
        product_info_dict["term"]  = list(map(int, product_info_dict["term"]))
        product_info_dict["productType"] = prod_packing[-1]
        return product_info_dict
    for i in range(1, sh.max_row + 1):
        col_dict = {
            "start": '',
            "end": ""
        }
        for j in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=i, column=j)
            # print(cell_obj.value, end=" ")
            # print("new val",((sh.cell(row =bold_indx+ 2, column = j)).font.b))
            if cell_obj.font.b and cell_obj.value!= None:
                bold_indx = i
                bold_flag =True
                # print("Bold===>",cell_obj.value, bold_indx)
                table_headings.append(copy.deepcopy(cell_obj.value))
            if i == bold_indx + 2 and bold_flag == True and ((sh.cell(row =bold_indx+ 2, column = j)).value)!=None and  ((sh.cell(row =bold_indx+ 2, column = j)).font.b) == False:
                # print(sh.cell(row=i, column=j).value,'real values')
                # print(i)
                row_dict["start"] = copy.deepcopy(i)
                col_list.append(copy.deepcopy(j))
                # print(col_list)
        if i>=bold_indx +2 and bold_flag == True and sh.cell(row=i, column=col_list[0]).value ==None:
            # print(sh.cell(row=i, column=col_list[0]) )
            row_dict["end"] = (i-1)
            col_dict["start"] = col_list[0]
            col_dict["end"] = col_list[-1]
            col_indxs.append(copy.deepcopy(col_dict))
            col_list = []
            row_indxs.append(copy.deepcopy(row_dict))
            row_dict={

            }
            col_dict= {

            }
            bold_flag =False
    table_headings.pop(0)

    table_headings.pop()


    row_indxs.pop()

    col_indxs.pop()

    print(row_indxs)
    print(col_indxs)
    print(table_headings)
    json_list = []
    main_json_list = []
    json_list_obj ={
        "version": version,
        "withEscrow":withEscrow,
        "provider":provider,
        "product":"",
        "values":[
        ]
    }

    data_value={
            }
    for index, value in enumerate(table_headings) :
        row_val = row_indxs[index]
        col_val = col_indxs[index]
        table_heading = table_headings[index]

        for row_range  in range(row_val['start'],row_val['end']+1):
            # print(table_heading)
            for col_range in range(col_val['start'],col_val['end']+1):
                min_amt=''
                if row_range > row_val['start']:
                    if table_heading =="FICO Adjustments":
                        json_list_obj["product"] = table_heading
                        min_max_val = find_min_max((sh.cell(row=row_range, column=col_val['start'])).value)
                        json_list_obj["minFico"] =copy.deepcopy(min_max_val["min"])
                        loan_amount = find_min_max((sh.cell(row=row_val["start"], column=col_range)).value)
                        # product_info = cal_term(table_heading)
                        # json_list_obj["product"] = table_heading
                        # json_list_obj["term"]= product_info["term"]
                        # json_list_obj["productType"] = product_info["productType"]
                        data_value["minLtv"] = copy.deepcopy(loan_amount["min"])
                        data_value["maxLtv"] = copy.deepcopy(loan_amount["max"])
                        data_value["minCltv"] = copy.deepcopy(loan_amount["min"])
                        data_value["maxCltv"] = copy.deepcopy(loan_amount["max"])

                        if row_range > row_val['start'] and col_range > col_val['start']:
                            cell_obj = sh.cell(row=row_range, column=col_range)
                            if cell_obj.value != None:
                                data_value["llpaValue"] = cell_obj.value
                            else:
                                pass
                            json_list_obj['values'].append(copy.deepcopy(data_value))
                            data_value = {

                            }
                        if min_max_val["max"] ==sys.maxsize:
                            json_list_obj["maxFico"] = copy.deepcopy(1000)
                        else:
                            json_list_obj["maxFico"] =copy.deepcopy(min_max_val["max"])
                    elif table_heading =="Loan Amount Adjustments":
                        json_list_obj["product"] = table_heading
                        min_max_val2 = find_min_max2((sh.cell(row=row_range, column=col_val['start'])).value)

                        json_list_obj['minLoan'] =copy.deepcopy(min_max_val2["min"])
                        json_list_obj['maxLoan'] =copy.deepcopy(min_max_val2["max"])
                        loan_amt = (sh.cell(row=row_val["start"], column=col_range)).value
                        loan_amount = find_min_max((sh.cell(row=row_val["start"], column=col_range)).value)
                        # product_info = cal_term(table_heading)
                        # json_list_obj["product"] = table_heading
                        # json_list_obj["term"]= product_info["term"]
                        # json_list_obj["productType"] = product_info["productType"]
                        data_value["minLtv"] = copy.deepcopy(loan_amount["min"])
                        data_value["maxLtv"] = copy.deepcopy(loan_amount["max"])
                        data_value["minCltv"] = copy.deepcopy(loan_amount["min"])
                        data_value["maxCltv"] = copy.deepcopy(loan_amount["max"])

                        if row_range > row_val['start'] and col_range > col_val['start']:
                            cell_obj = sh.cell(row=row_range, column=col_range)
                            if cell_obj.value != None:
                                data_value["llpaValue"] = cell_obj.value
                            else:
                                pass
                            json_list_obj['values'].append(copy.deepcopy(data_value))
                            data_value = {

                            }

                    elif table_heading =="Purpose / Property Adjustments":
                        json_list_obj["product"] = table_heading
                        propertyType =(sh.cell(row=row_range, column=col_val['start'])).value
                        json_list_obj["propertyType"] =propertyType
                        loan_amt=(sh.cell(row=row_val["start"], column=col_range)).value
                        loan_amount= find_min_max((sh.cell(row=row_val["start"], column=col_range)).value)
                    # product_info = cal_term(table_heading)
                    # json_list_obj["product"] = table_heading
                    # json_list_obj["term"]= product_info["term"]
                    # json_list_obj["productType"] = product_info["productType"]
                        data_value["minLtv"]=copy.deepcopy(loan_amount["min"])
                        data_value["maxLtv"]=copy.deepcopy(loan_amount["max"])
                        data_value["minCltv"]=copy.deepcopy(loan_amount["min"])
                        data_value["maxCltv"]=copy.deepcopy(loan_amount["max"])

                        if row_range > row_val['start'] and col_range > col_val['start']:
                            cell_obj = sh.cell(row=row_range, column=col_range)
                            if cell_obj.value != None:
                                data_value["llpaValue"] = cell_obj.value
                            else:
                                pass
                            json_list_obj['values'].append(copy.deepcopy(data_value))
                            data_value = {

                            }
                    elif table_heading =="Other Adjustments":
                        json_list_obj["product"] = table_heading
                        print(row_val['start']-1)

                        if row_range >= row_val['start']-2:
                            adjustmentType =(sh.cell(row=row_range-1, column=col_val['start'])).value
                            print(adjustmentType)
                            json_list_obj["otherAdjustmentType"] = adjustmentType
                        # main_json_list.append(copy.deepcopy(json_list_obj))

                        # isPerDay = False
                        # llpaVal =''
                        # col_llpa_val = ''
                        #
                        # if row_range > row_val['start']-1 and col_range > col_val['start']:
                        #     cell_obj = sh.cell(row=row_range, column=col_range)
                        #     col_llpa_val = cell_obj.value
                        #     print(col_llpa_val)
                        #
                        #     if cell_obj.value != None:
                        #         if ' / day' in str(col_llpa_val):
                        #             # json_list_obj["otherAdjustmentType"] = copy.deepcopy( (sh.cell(row=row_range, column=col_val['start'])).value)
                        #             isPerDay = True
                        #             json_list_obj["otherAdjustmentType"] = (
                        #                 (sh.cell(row=row_range, column=col_val['start'])).value)
                        #             llpaVal, val = col_llpa_val.split(' / day')
                        #             llpaVal = llpaVal.replace('(', '-')
                        #             llpaVal = llpaVal.replace(')', '')
                        #             data_value["llpaValue"] = llpaVal
                        #             llpaVal = col_llpa_val
                        #             data_value["llpaValue"] = col_llpa_val
                        #             data_value["isPerDay"] = isPerDay
                        #
                        #
                        #
                        #         elif '/ day' not in str(col_llpa_val):
                        #             json_list_obj["otherAdjustmentType"] = (
                        #                 (sh.cell(row=row_range, column=col_val['start'])).value)
                        #             print(col_llpa_val)
                        #             isPerDay = False
                        #             llpaVal = col_llpa_val
                        #             data_value["llpaValue"] = col_llpa_val
                        #             data_value["isPerDay"] = isPerDay
                        #
                        #         # data_value["llpaValue"] = cell_obj.value
                        #     else:
                        #         pass
                        #     json_list_obj['values'].append(copy.deepcopy(data_value))
                        #     data_value = {
                        #
                        #     }

                    ##Second
                        # if col_range > col_val['start']:
                        #     col_llpa_val = (sh.cell(row=row_val["start"]-1, column=col_range)).value
                        # print(col_llpa_val)
                        #
                        # if ' / day' in str(col_llpa_val):
                        #     # json_list_obj["otherAdjustmentType"] = copy.deepcopy( (sh.cell(row=row_range, column=col_val['start'])).value)
                        #     isPerDay = True
                        #     json_list_obj["otherAdjustmentType"] = copy.deepcopy(
                        #         (sh.cell(row=row_range, column=col_val['start'])).value)
                        #     llpaVal, val = col_llpa_val.split(' / day')
                        #     llpaVal = llpaVal.replace('(','-')
                        #     llpaVal = llpaVal.replace(')','')
                        #
                        # elif '/ day' not in str(col_llpa_val):
                        #     print(col_llpa_val)
                        #     isPerDay =False
                        #     llpaVal = col_llpa_val
                        #
                        #
                        #


            if row_range > row_val['start'] and table_heading !="Other Adjustments":
                main_json_list.append(copy.deepcopy(json_list_obj))
                json_list_obj = {
                                "version": version,
                                "withEscrow": withEscrow,
                                 "provider": provider,
                                "product":"",
                                "values":[
                                ]
                        }
            # if row_range > row_val['start']-2 and table_heading == "Other Adjustments":
            #     main_json_list.append(copy.deepcopy(json_list_obj))
            #     json_list_obj = {
            #             "version": version,
            #             "withEscrow": withEscrow,
            #             "provider": provider,
            #             "product": "",
            #             "values": [
            #             ]
            #         }
    print("here is what u looking for length of the array isn't it...!" ,len(main_json_list))
    json_object = json.dumps(main_json_list, indent=4)
    with open("PortfolioExpressLLPAs.json","w") as j_file:
        print("JSON file is generated Happppilyy...!")
        j_file.write(json_object)
    return (main_json_list)
    #
    #
    #
